#coding=<UTF-8>
from openpyxl import load_workbook

from Board import *
from appjar import gui
import logging



class BS:

	def __init__(self, file, param):
		self.win = None
		self.boards = []
		self.input = {} # will contain the input from the user
		self.regions = []   # list of all the regions
		self.idies = []		 # list of all the idies
		self.file = file	   # the excel file with the board details
		self.param = param  # the parameter to multiply all the size's
		self.A3 = {"width": str(self.convertX(30)), "high": str(self.convertX(42))}
		self.exitPressed = False  # used for 'remaindSaving' function to avoid double quiting

	def start(self):
		self.win = gui("board system", "500x400")
		self.update(self.file)  # read the boards from excel and the posts from "posts.txt". update boards, regions and idies
		self.win.addLabelAutoEntry("board id:", self.idies, row=1)
		self.win.setEntryDefault("board id:", "If empty, action will aplly to all boards/region.")
		self.win.addLabelAutoEntry("region:", self.regions, row=5)
		self.win.setEntryDefault("region:", "If region and id field are empty, action will aplly to all boards.")
		self.win.addLabelEntry("post's name:")
		self.win.addLabelEntry("post's width:")
		self.win.setEntryDefault("post's width:", "If width and high field are empty, the size will be A3.")
		self.win.addLabelEntry("post's high:")
		self.win.setEntryDefault("post's high:", "If high and width field are empty, the size will be A3.")
		self.win.addButton("add post automatic", self.buttonAddPostAutomatic)
		self.win.addButton("add post", self.buttonAddPost)
		self.win.addButton("show board", self.buttonShowBoard)
		self.win.addButton("remove post", self.buttonRemovePost)
		self.win.addButton("remove all posts", self.buttonRemoveAllPosts)
		self.win.addButton("save", self.save)
		self.win.setStopFunction(self.remindSaving)
		self.win.go()

	# convert the size to wanted size
	def convertX(self, x):
		if isinstance(x, str):
			if x == "":
				return x
			return str(int(x) * self.param)
		return int(x) * self.param

	# convert the size back when saving the posts
	def convertBack(self, x):
		if isinstance(x, str):
			if x == "":
				return x
			return str(int(x) / self.param)
		return int(x) / self.param

	# remind the user to save
	def remindSaving(self):
		if not self.exitPressed:
			self.exitPressed = True
			toQuit = self.win.yesNoBox("remined saving", "do you remembered to save?")
			if toQuit:
				return True
			self.exitPressed = False
			return False
		return False

	def isIdEntered(self, error=True):
		if self.input["id"] == "":
			if error:
				self.win.errorBox("error","Enter id board")
			return False
		return True

	# check if the board with 'id' exist in self.boards
	def isBoardExist(self, id):
		board = self.findBoard(id)
		if board == None:
			self.win.errorBox("error","board number " + id + " dosent exist")
			return False
		return True

	# check if the region exist in self.region
	def isRegionExist(self, region):
		for board in self.boards:
			if board.region == region:
				return True
		self.win.errorBox("error","region: \""+ region +"\" dosent exist")
		return False


	def buttonAddPostAutomatic(self):
		# pdb.set_trace()
		self.input = self.getInput()
		input = self.input
		if input is None:
			return
		if self.isPostDetalisEntere():
			if self.isIdEntered(error=False):
				if self.isBoardExist(input["id"]):
					board = self.findBoard(input["id"])
					id = board.addPostAuto(input["name"], [int(input["width"]), int(input["high"])])
					if not id: # id cant add id=None. else id is the target board
						self.win.errorBox("error", "cant add the post to board number: " + board.id)
					else:
						self.win.infoBox("information", "post name: \"" + input["name"] + "\" added to board number " + id)
			else:
				if input["region"] == "":
					for b in self.boards:
						if not b.addPostAuto(input["name"], [int(input["width"]), int(input["high"])]):
							self.win.errorBox("error","cant add post name: \"" + input["name"] + "\" to board number: " + b.id)
					self.win.infoBox("information","post name: \"" + input["name"] + "\" added  to all board")
				else:
					if self.isRegionExist(input["region"]):
						for b in self.boards:
							if b.region == input["region"]:
								if not b.addPostAuto(input["name"], [int(input["width"]), int(input["high"])]):
									self.win.errorBox("error","cant add post name: \"" + input["name"] +
													  "\" to board number: " + b.id)
					self.win.infoBox("information", "post name: \"" + input["name"] + "\" added to region: \"" + input["region"] + "\"")

	def buttonRemoveAllPosts(self):
		self.input = self.getInput()
		input = self.input
		if input is None:
			return
		if self.isIdEntered(error=False):
			if self.isBoardExist(input["id"]):
				if self.win.okBox("ask","remove all posts from board number: " + input["id"] + "?"):
					board = self.findBoard(input["id"])
					board.clear()
					self.win.infoBox("information","removed all posts from board number " + input["id"] + ".")
					return
		elif input["region"] == "":
			if self.win.okBox("ask","removed all posts from all the boards?"):
				for board in self.boards:
					board.clear()
				self.win.infoBox("information","all posts removed from all boards.")
		elif self.isRegionExist(input["region"]):
				if self.win.okBox("ask","remove all posts from: \"" + input["region"] + "?"):
					for board in self.boards:
						if board.region == input["region"]:
							board.clear()
					self.win.infoBox("information","all posts removed from: \"" + input["region"] + "\".")

	def buttonRemovePost(self):
		self.input = self.getInput()
		input = self.input
		if input is None:
			return
		if self.isNameEntered():
			if self.isIdEntered(error=False):
				if self.isBoardExist(input["id"]):
					if self.win.okBox("ask","remove post name: \"" + input["name"] + "\" from board number: " + input["id"] + "?"):
						board = self.findBoard(input["id"])
						post = board.findPost(input["name"])
						board.removePost(post)
						self.win.infoBox("information",
							"post name: \"" + input["name"] + "\", removed from board number " + input["id"])
			else:
				if input["region"] == "":
					if self.win.okBox("ask","removed post name: \"" + input["name"] + "\" from all the boards?"):
						for board in self.boards:
							board.removePost(board.findPost(input["id"]))
							self.win.infoBox("information","post name: \"" + input["name"] + "\", rermoved from all the boards")
				else:
					if self.isRegionExist(input["region"]):
						if self.win.okBox("ask","remove post name: \"" + input["name"] + "\" from: \"" + input["region"] + "\"?"):
							for board in self.boards:
								if board.region == input["region"]:
									board.removePost(board.findPost(input["id"]))
							self.win.infoBox("information","post name: \"" + input["name"] + "\", rermoved from \""
											 + input["region"] + "\"")
	def isNameEntered(self):
		if self.input["name"] == "":
			self.win.errorBox("error","Enter post's name")
			return False
		return True

	def buttonShowBoard(self):
		self.input = self.getInput()
		input = self.input
		if input is None:
			return
		if self.isIdEntered(error=False):
			if self.isBoardExist(input["id"]):
				board = self.findBoard(input["id"])
				board.init()
				board.show()
		else:
			if input["region"] == "":
				for board in self.boards:
					board.init()
					board.show()
					if checkForkey("q"): # if the user pressed shift + q, return to the main board
						break
			else:
				if self.isRegionExist(input["region"]):
					for board in self.boards:
						if board.region == input["region"]:
							board.init()
							board.show()
							if checkForkey("q"):  # if the user pressed shift + q, return to the main board
								break

	def buttonAddPost(self):
		self.input = self.getInput()
		input = self.input
		if input is None:
			return
		if self.isPostDetalisEntere():
			if self.isIdEntered(error=False):
				if self.isBoardExist(input["id"]):
					board = self.findBoard(input["id"])
					board.init()
					board.UserAddPost(input["name"], [int(input["width"]), int(input["high"])])
			else:
				if input["region"] == "":
					for board in self.boards:
						board.init()
						board.UserAddPost(input["name"], [int(input["width"]), int(input["high"])])
						if checkForkey("q"):  # if the user pressed shift + q, return to the main board
							break
				else:
					if self.isRegionExist(input["region"]):
						for board in self.boards:
							if board.region == input["region"]:
								board.init()
								board.UserAddPost(input["name"], [int(input["width"]), int(input["high"])])
								if checkForkey("q"):  # if the user pressed shift + q, return to the main board
									break

	# check if the detail needed to add post is enterd.
	def isPostDetalisEntere(self):
		input = self.input
		if not self.isNameEntered():
			return False
		if input["high"] == "" and input["width"] == "":
			input["width"] = self.A3["width"]
			input["high"] = self.A3["high"]
			return True
		if input["high"] == "":
			self.win.errorBox("error","Enter post's high")
			return False
		if input["width"] == "":
			self.win.errorBox("error","Enter post's width")
			return False
		return True

	def getInput(self):
		try:
			return {
				"id": self.win.getEntry("board id:"),
				"name": self.win.getEntry("post's name:"),
				"width": self.convertX(self.win.getEntry("post's width:")),
				"high": self.convertX(self.win.getEntry("post's high:")),
				"region": self.win.getEntry("region:")
			}
		except ValueError as v:
			self.printError("the width and high field cannot contain letters.",v,quit=False)
			return None


	def findBoard(self, id):
		for board in self.boards:
			if board.id == id:
				return board
		else:
			return None

	def addBoard(self, board):
		self.boards.append(board)
		self.regions.append(board.region)
		self.idies.append(board.id)

	def addBoards(self, boards):
		for board in boards:
			self.addBoard(board)

	def addBoardByDetail(self, width, high, id):
		board = Board(width, high, id)
		self.addBoard(board)

	def removeBoard(self, board):
		self.boards.remove(board)
		self.idies.remove(board.id)

	def removeBoards(self, boards):
		for board in boards:
			self.removeBoard(board)

	# save the details of the post in file "posts.txt"
	def save(self):
		file = openFile("posts.txt", "w")
		for board in self.boards:
			file.write(board.id + "\n")
			for post in board.posts:
				file.write(post.name + "\n" + str(self.convertBack(post.location[0])) + "\n"
						   + str(self.convertBack(post.location[1])) \
						   + "\n" + str(self.convertBack(post.size[0])) + "\n"
						   + str(self.convertBack(post.size[1])) + "\n")
			file.write("\n")
		file.close()
		self.win.infoBox("information", "saved succefuly")

	# read the posts and the boards details from "posts.txt" and the given file. also update the regions from the file.
	def update(self,file):
		postsForEachBoard = self.getPostsFromFile()
		boards, regions = self.getBoardsDetailsAndRegion(file)
		boardsToRemove = []
		for board in boards:
			if board.id in postsForEachBoard.keys():
				board.addPosts(postsForEachBoard[board.id])
			if not board.id.isdigit() and board.id.endswith("a"):
				for child in boards:
					if child.id.startswith(board.id[0:len(board.id)-1]) and not child.id.endswith("a"):
						board.children.append(child)
						boardsToRemove.append(child)
			self.idies.append(board.id)
		for child in boardsToRemove:
			boards.remove(child)
		self.boards = boards
		self.regions = regions

	# read the details of the posts from "posts.txt". after each boards id there is the posts belongs to him,
	# and in the end a blank line.
	# post detail is name, x, y, width, high. in seperate lines.
	def getPostsFromFile(self):
		# pdb.set_trace()
		boards = {}	 # will contain the id in the key and the posts in the value
		linesForPost = 5  # number of lines for each post
		indexForBoard = 0   # indexing which board number we are now
		try:
			file = openFile("posts.txt", "r").readlines()
		except FileNotFoundError:
			openFile("posts.txt", "w+").close()
		file = openFile("posts.txt", "r").readlines()
		while indexForBoard < len(file):
			posts = []  # will contaion the posts that belongs to the board
			id = file[indexForBoard][0:len(file[indexForBoard]) - 1]  # save the id off the board
			index = indexForBoard + 1  # start reading the posts off the board
			while file[index] is not "\n" and index + 2 < len(file):
				posts.append(
					Post(file[index][0:len(file[index]) - 1], [self.convertX(float(file[index + 1][0:len(file[index + 1]) - 1])),
															   self.convertX(float(file[index + 2][0:len(file[index + 2]) - 1]))],
													[self.convertX(float(file[index + 3][0:len(file[index + 3]) - 1])),
													self.convertX(float(file[index + 4][0:len(file[index + 4]) - 1]))]))
				index += linesForPost   # nove to the next post
				indexForBoard += linesForPost   # advance the index until the next board
			boards[id] = posts
			indexForBoard += 2  # skip the blank line
		return boards

	# for case of sizes in meter
	def converToCantimters(self,width, high):
		if width > 10:
			return int(width), int(high)
		return int(width * 100), int(high * 100)

	# get the width and high from the B column. except number, then some char then a number.
	def getWidthHige(self,cellData):
		index = 0
		for c in cellData:
			if not c.isdigit() and not c == '.':
				break
			index += 1
		width, high = self.converToCantimters(float(cellData[index + 1: len(cellData)]), float(cellData[0: index]))
		return self.convertX(width), self.convertX(high)

	# read the details of the boards and their region from excel file. the sheet name is the region, column A for id
	# and column B for width and high.
	def getBoardsDetailsAndRegion(self, file):
		try:
			wb = load_workbook(findPath(file))
		except FileNotFoundError:
			self.printError("file name: \"" + file + "\" has not found in the current directory.",FileNotFoundError)
		regions = wb.sheetnames
		boards = []
		for region in regions:
			sheet = wb[region]
			row = 2
			columns = ["A", "B"]
			while sheet["A" + str(row)].value:
				try:
					width, high = self.getWidthHige(sheet["B" + str(row)].value)
				except ValueError:
					self.printError("error in region " + region +" in row " + str(row) + ". \ncolumn B suppose to be in the next foramt:\n"
							"number, char to seprate, and number. for example: 80/90 or 120x100.",ValueError)	
				boards.append(Board(width, high, sheet["A" + str(row)].value, region))
				row += 1
		return boards, regions

	def printError(self,msg,errorInst,toQuit=True):
		self.win.errorBox("error",msg)
		logError(errorInst)
		if toQuit:
			quit()
		
if __name__ == "__main__":
	logging.basicConfig(filename='log.log', level=logging.DEBUG, filemode='w')
	bs = BS("boards.xlsx",3)
	bs.start()


